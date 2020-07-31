from bs4 import BeautifulSoup as soup
import pandas as pd
import openpyxl
import requests
import sys
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
import time
import os



# Below code is for Sabine Pass - Kinder Morgan KMLP
#print("Below are volumes for KMLP - Sabine Pass (Kinder Morgan)")
spkmlp_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Path must be specified depending on local machine
spkmlp_driver.get('https://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code=KMLP') # get request for Selenium to download the page in a proxy browser
spkmlp_retrieve_xpath = '//*[@id="WebSplitter1_tmpl1_ContentPlaceHolder1_HeaderBTN1_btnRetrieve"]' # ignore - xpath for retrieve button - debugging
spkmlp_retrieve_click = spkmlp_driver.find_element_by_name('ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnRetrieve').click() # automatically initiates a click on retrieve button to download content
spkmlp_tsq = spkmlp_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[4]/div[1]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[16]/td/div/table/tbody/tr/td[8]").get_property("innerText") # retrieves the TSQ values for KMLP Liquefaction Deliveries
spkmlp_driver.close() # closes opened proxy page



# Below is script for Sabine Pass - TGT Lighthouse Road M4662 MP 13
#print("Below are volumes for TGT Lighthouse Road - Sabine Pass")
sbtgt_url = "http://www.1line.williams.com/ebbCode/OACQueryRequest.jsp?BUID=80&type=OAC" # URL
sbtgt_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
sbtgt_driver.get(sbtgt_url) # Driver proxy request
sbtgt_driver.find_element_by_css_selector("body > form > p > input[type=submit]").click() # retrieve button click
WebDriverWait(sbtgt_driver, 30).until(EC.number_of_windows_to_be(2)) # Driver wait request to prevent a timeout error
sbtgt_report = sbtgt_driver.window_handles[1] # Prioritizes the result window to initiate scrape
sbtgt_driver.switch_to_window(sbtgt_report) # Switches to window to enable scrape
#sbtgt_driver.switch_to_window("OACreport")
sbtgt_lighthouse = sbtgt_driver.find_element_by_xpath("/html/body/form/table/tbody/tr[2]/td/div/table[3]/tbody/tr[223]/td[9]").get_property("textContent") # Selects the TSQ and retrieves TSQ text
#print(sbtgt_lighthouse)
sbtgt_driver.close() # Closes result page
sbtgt_driver.quit() # Terminates proxy browser


# Below is the script for Sabine Pass - Creole Trail Cheniere LNG Connection
#print("Below are volumes for Creole Trail - Sabine Pass (LNG Connection)")
ct_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # filepath for driver
ct_driver.get('https://lngconnection.cheniere.com/#/ctpl') # Opens proxy browser and fetches URL
ct_elems = WebDriverWait(ct_driver, 30).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="capacityGrid"]/table/tbody/tr[1]/td[4]'))) # Wait request to prevent timeout error
ct_driver.find_element_by_xpath("/html/body/app/main/side-menu/nav/ul/li[2]/a").click() # Clicks on Capacity
ct_driver.find_element_by_xpath("/html/body/app/main/side-menu/nav/ul/li[2]/ul/li[1]/a").click() # Clicks on Operationally Available
ct_table_elems = WebDriverWait(ct_driver, 30).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="gridOperationallyAvailable"]/table/tbody/tr[1]/td[5]'))) # Wait request to finish loading content
#ct_html = ct_driver.page_source

ct_trunkline_ctpl = ct_driver.find_element_by_xpath('//*[@id="gridOperationallyAvailable"]/table/tbody/tr[3]/td[14]').get_property('innerText') # TSQ
ct_transco_ctpl = ct_driver.find_element_by_xpath('//*[@id="gridOperationallyAvailable"]/table/tbody/tr[2]/td[14]').get_property('innerText') # TSQ
ct_tetco_ctpl = ct_driver.find_element_by_xpath('//*[@id="gridOperationallyAvailable"]/table/tbody/tr[1]/td[14]').get_property('innerText') # TSQ
ct_spliq_ctpl = ct_driver.find_element_by_xpath('//*[@id="gridOperationallyAvailable"]/table/tbody/tr[4]/td[14]').get_property('innerText') # TSQ
ct_driver.close()



time.sleep(2)
# Below is the script for Corpus Christi - LNG Connection
#print("Below are volumes for Corpus Christi - LNG Connection")
cc_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
cc_driver.get('https://lngconnection.cheniere.com/#/ccpl') # Opens proxy page with the URL
cc_elems = WebDriverWait(cc_driver, 10).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="capacityGrid"]/table/tbody/tr[1]/td[4]'))) # Wait request for content to load
#cc_html = cc_driver.page_source

time.sleep(2)
cc_tgp_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[2]/td[4]').get_property('innerText') # TSQ
cc_transco_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[3]/td[4]').get_property('innerText') # TSQ
cc_ngpl_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[4]/td[4]').get_property('innerText') # TSQ
cc_kmtejas_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[6]/td[4]').get_property('innerText') # TSQ
cc_epdtexas_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[5]/td[4]').get_property('innerText') # TSQ
cc_ccliq = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[7]/td[4]').get_property('innerText') # TSQ
cc_driver.close()


# Below is the script for Cameron LNG

ci_url = "http://www.gasnom.com/ip/cameron/oauc.cfm?type=1" # URL for Gasnom
ci_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for driver
ci_driver.get(ci_url) # Opens proxy page for the URL
ci_texeastcip = ci_driver.find_element_by_xpath("/html/body/div[1]/div[3]/table[2]/tbody/tr[13]/td[9]").get_property("innerText") # TSQ
ci_texeastern = ci_driver.find_element_by_xpath("/html/body/div[1]/div[3]/table[2]/tbody/tr[15]/td[9]").get_property("innerText") # TSQ
ci_tenngaspipe = ci_driver.find_element_by_xpath("/html/body/div[1]/div[3]/table[2]/tbody/tr[10]/td[9]").get_property("innerText") # TSQ
ci_termdeliv = ci_driver.find_element_by_xpath("/html/body/div[1]/div[3]/table[2]/tbody/tr[3]/td[9]").get_property("innerText") # TSQ
ci_driver.close() # Closes proxy page


# Below is the volume for CGT Cameron Access

cgt_url = "http://www.columbiapipeinfo.com/infopost/ReportViewer.aspx?/InfoPost/OperationallyAvailableCapacity&pAssetNbr=51" # URL
cgt_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
cgt_driver.get(cgt_url) # Opens proxy page with URL
cgt_elems = WebDriverWait(cgt_driver, 10).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr[4]/td[3]/div/div[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[8]/td[3]/table/tbody/tr[3]/td[1]/div/div'))) # Waits for content to finish loading to prevent timeout error
cgt_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[1]/td[2]/div/select").click() # Clicks on asset
cgt_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[1]/td[2]/div/select/option[1]").click() # Selects CGT Asset
cgt_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[3]/table/tbody/tr/td/input").submit() # Clicks on view report and submits form
cgt_findwait = WebDriverWait(cgt_driver, 20).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[3]/td[2]/div/select'))) # Wait request for content to load to prevent timeout error
cgt_findwait.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[3]/td[2]/div/select").click() # Clicks on location
cgt_findwait.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[3]/td[2]/div/select/option[20]").click() # Clicks on Cameron Pipeline (loc: 4246)
cgt_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[3]/table/tbody/tr/td/input").click() # Clicks on view report
cgt_delivwait = WebDriverWait(cgt_driver, 20).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[2]/td[5]/div/select'))) # Wait request for flow indicator to be clickable
cgt_delivwait.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[2]/td[5]/div/select').click() # clicks on flow indicator
cgt_delivwait.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[2]/td[5]/div/select/option[2]').click() # selects 'delivery' points
WebDriverWait(cgt_driver, 20).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[3]/table/tbody/tr/td/input')))
cgt_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[3]/table/tbody/tr/td/input').click()

#cgt_tsq = cgt_driver.find_element_by_css_selector('#P9646b5f1df1245a8b40443487b9302b1_2_129iT0R0x0').get_property('innerText')
#print(cgt_tsq)

cgt_source = cgt_driver.page_source
#print(cgt_source)
cgt_driver.close()

forsoup = cgt_source

#print(etree.tostring(root, encoding='unicode',pretty_print=True))

s = soup(forsoup,'html.parser')
c = s.find_all('div',attrs={'style':'width:23.06mm;min-width: 23.06mm;'})[1]
cgt_tsq = int(str(c).split('>')[1].split('<')[0].replace(",",""))
#print(cgt_tsq)


# Below is the script for Elba - Elba Express Connector (ELC Chatham Deliveries) - ELC/EEC EEC TO ELC CHATHAM 660700 - Kinder Morgan
#print("Below are volumes for Elba Express Chatham Deliveries - Kinder Morgan")
elba_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Path must be specified depending on local machine
elba_driver.get('https://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code=EEC') # get request for Selenium to download the page in a proxy browser
elba_driver.find_element_by_name("ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnRetrieve").click() # Clicks on retrieve button
elba_elems = WebDriverWait(elba_driver, 10).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[4]/div[1]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[3]/td/div/table/tbody/tr/td[8]'))) # Wait request to finish loading content
elba_tsq = elba_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[4]/div[1]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[3]/td/div/table/tbody/tr/td[8]").get_property("innerText") # captures TSQ of point specified
#print(elba_tsq)
elba_driver.close() # Closes proxy page

# Print commands for all variables with TSQ values stored
print("Below are volumes for Sabine Pass KMLP")
print(spkmlp_tsq)
print("Below are volumes for TGT Lighthouse")
print(sbtgt_lighthouse)
print("Below are volumes for Sabine Pass - Creole Trail")
print(ct_trunkline_ctpl)
print(ct_transco_ctpl)
print(ct_tetco_ctpl)
print(ct_spliq_ctpl)
print("Below are volumes for Corpus Christi")
print(cc_tgp_ccpl)
print(cc_transco_ccpl)
print(cc_ngpl_ccpl)
print(cc_kmtejas_ccpl)
print(cc_epdtexas_ccpl)
print(cc_ccliq)
print("Below are volumes for Cameron Interstate")
print(ci_texeastcip)
print(ci_texeastern)
print(ci_tenngaspipe)
print(ci_termdeliv)
print("Below is the volume for CGT Cameron Access")
print(cgt_tsq)
print("Below are volumes for Elba - Chatham Deliveries")
print(elba_tsq)
print("Below are excel downloaded tsq's")

# Cove Point File Download
cove_url = 'https://dekaflow.dominionenergy.com/servlet/InfoPostServlet?region=null&company=cpt&method=headers&category=Capacity&subcategory=Operationally+Available'
cove_options = webdriver.ChromeOptions() # opens WebDriver
cove_dpath = '/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/cove burner' # filepath for WebDriver
os.makedirs(cove_dpath) # Makes directory
#options.add_argument("download.default_directory=/Users/rajgarkhedkar/Desktop/Enkon/2020")
cove_prefs = {'download.default_directory' : '/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/cove burner'} # sets proxy browser preferences to download excel file to burner directory
cove_options.add_experimental_option('prefs', cove_prefs) # sets preferences for download


cove_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver',options=cove_options) # Driver with filepath and directory preference set
cove_driver.get(cove_url) # opens proxy browser with the cove point url

cove_download = cove_driver.find_element_by_xpath('/html/body/center/table/tbody/tr[7]/td[5]/a').click() # clicks on download excel file

#waits for the selenium to complete download. If false, keeps the loop open until true
def download_begin(cove_driver):
            if len(os.listdir(cove_dpath)) == 0:
                time.sleep(0.5)
                return False
            else:
                return True
WebDriverWait(cove_driver, 120).until(download_begin)
cove_burnfile=os.listdir('/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/cove burner')[0] # selects file from burner directory
cove_burnfile = cove_dpath+'/'+cove_burnfile # adds a slash to format the name properly

pd.set_option('display.max_columns',1000)
pd.set_option('display.max_rows',1000)
cove_dfs = pd.read_csv(cove_burnfile) # reads csv into DataFrame
cp_trans_pleas_tsq = int(cove_dfs[(cove_dfs['Loc Name']=='TRANSCO PLEASANT VALLEY') & (cove_dfs['Loc Purp Desc']=='Receipt Location')]['Total Scheduled Qty'].values[0]) # specific TSQ
cp_col_lou_tsq = int(cove_dfs[(cove_dfs['Loc Name']=='COLUMBIA LOUDOUN') & (cove_dfs['Loc Purp Desc']=='Receipt Location')]['Total Scheduled Qty'].values[0]) # specific TSQ
cp_cove_poi_tsq = int(cove_dfs[(cove_dfs['Loc Name']=='COVE POINT PLANT') & (cove_dfs['Loc Purp Desc']=='Delivery Location')]['Total Scheduled Qty'].values[0]) # specific TSQ
#print(cove_dfs) # prints dataframe
cove_driver.close() # closes proxy browser

print(cp_trans_pleas_tsq)
print(cp_col_lou_tsq)
print(cp_cove_poi_tsq)

os.remove(cove_burnfile) # removes burner csv
os.rmdir(cove_dpath) # removes burner directory


# Gulf South File Download
gulf_url = 'https://infopost.bwpipelines.com/Posting/default.aspx?Mode=Display&Id=11&tspid=1'

gulf_options = webdriver.ChromeOptions() # enables options preferences
gulf_dpath = '/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/gulf burner' # filepath for burner directory
os.makedirs(gulf_dpath) # makes burner directory
gulf_prefs = {'download.default_directory' : '/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/gulf burner'} # sets download preferences to burner directory
gulf_options.add_experimental_option('prefs', gulf_prefs) # adds download preferences to webdriver


#gulf_url = 'https://infopost.bwpipelines.com/Posting/default.aspx?Mode=Display&Id=11&tspid=1'
gulf_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver',options=gulf_options) # initializes WebDriver with set download preferences
gulf_driver.get(gulf_url) # opens url

gulf_download = gulf_driver.find_element_by_xpath('/html/body/form/div[3]/div/div[2]/div[3]/table/tbody/tr[8]/td[3]/a').click() # clicks download on excel file

# Function to ensure download completes
def download_begin(gulf_driver):
            if len(os.listdir(gulf_dpath)) == 0:
                time.sleep(0.5)
                return False
            else:
                return True
WebDriverWait(gulf_driver, 120).until(download_begin)
gulf_burnfile=os.listdir('/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/gulf burner')[0] # chooses file from directory
gulf_burnfile = gulf_dpath+'/'+gulf_burnfile # adds slash to format properly

pd.set_option('display.max_columns',1000)
pd.set_option('display.max_rows',1000)
gulf_dfs = pd.read_excel(gulf_burnfile) # reads excel file into DataFrame
#print(gulf_dfs) # prints DataFrame
gulf_south_tsq=int(gulf_dfs[gulf_dfs['Loc']==23700]['Total Scheduled Qty'].values[0]) # specific TSQ
gulf_driver.close() # closes driver

print(gulf_south_tsq) # prints specific loc

os.remove(gulf_burnfile) # removes burner excel file
os.rmdir(gulf_dpath) # removes directory


# Last TETCO
# https://infopost.spectraenergy.com/infopost/TEHome.asp?Pipe=TE
tet_url = 'https://rtba.spectraenergy.com/InformationalPosting/Default.aspx?bu=TE&Type=OA'

tet_options = webdriver.ChromeOptions() # initializes options for proxy browser
tet_dpath = '/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/tetco burner' # sets a burner directory
os.makedirs(tet_dpath) # creates directory
tet_prefs = {'download.default_directory' : '/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/tetco burner'} # creates preferences for file download
tet_options.add_experimental_option('prefs', tet_prefs) # sets preferences


#gulf_url = 'https://infopost.bwpipelines.com/Posting/default.aspx?Mode=Display&Id=11&tspid=1'
tet_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver',options=tet_options) # initializes WebDriver with set options
tet_driver.get(tet_url) # openx proxy browser on gulf south page


tet_driver.find_element_by_xpath('/html/body/form/div[3]/div[1]/div[2]/div[2]/div[2]/div/div[2]/select').click() # mouseclick for download action on page
tet_driver.find_element_by_xpath('/html/body/form/div[3]/div[1]/div[2]/div[2]/div[2]/div/div[2]/select/option[1]').click() # mouseclick for download action on page
tet_download = tet_driver.find_element_by_xpath('/html/body/form/div[3]/div[1]/div[2]/div[2]/div[2]/div/div[3]/a[2]').click() # clicks download

# function to ensure file has downloaded
def download_begin(tet_driver):
            if len(os.listdir(tet_dpath)) == 0:
                time.sleep(0.5)
                return False
            else:
                return True
WebDriverWait(tet_driver, 120).until(download_begin)
tet_burnfile=os.listdir('/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/tetco burner')[0] # selects file from directory
tet_burnfile = tet_dpath + '/' + tet_burnfile # adds slash to format properly

pd.set_option('display.max_columns',1000)
pd.set_option('display.max_rows',1000)
tet_dfs = pd.read_csv(tet_burnfile) # reads csv into DataFrame
last_tetco_tsq=int(tet_dfs[tet_dfs['Loc']==73912]['Total_Scheduled_Quantity'].values[0]) # specific tsq
#print(tet_dfs) # prints DataFrame
tet_driver.close() # closes proxy browser

print(last_tetco_tsq)

os.remove(tet_burnfile) # removes burner csv
os.rmdir(tet_dpath) # removes burner directory



# NGPL East - Error solved

ngpleast_url = 'https://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code=NGPL'

ngpleast_options = webdriver.ChromeOptions() # initializes options
ngpleast_dpath = '/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/ngpleast burner' # sets burner directory
os.makedirs(ngpleast_dpath) # creates burner directory
ngpleast_prefs = {'download.default_directory' : '/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/ngpleast burner'} # initializes preferences to download to burner directory
ngpleast_options.add_experimental_option('prefs', ngpleast_prefs) # sets options

ngpleast_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver',options=ngpleast_options) # initializes WebDriver with download preferences
ngpleast_driver.get(ngpleast_url) # opens proxy browser to NGPL page
ngpleast_driver.implicitly_wait(5) # wait for 5 seconds

ngpleast_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[1]/table/tbody/tr/td[2]/table/tbody/tr/td/input[2]').click() # clicks on download excel file

# function to make sure big excel file is downloaded
def download_begin(ngpleast_driver):
            if len(os.listdir()) == 0:
                time.sleep(0.5)
                return False
            else:
                return True
WebDriverWait(ngpleast_driver, 120).until(download_begin)
ngpleast_driver.close() # closes proxy browser
time.sleep(5) # pauses script for 5 seconds to allow file to be saved properly to ensure no troubles reading

ngpleast_burnfile_k=os.listdir(ngpleast_dpath)[0] # selects file from directory
ngpleast_burnfile = ngpleast_dpath+'/'+ngpleast_burnfile_k # adds slash to format properly

'''
# Below if-else statement is in case file has .crdownload at the end of name. Will not read into DataFrame properly if .crdownload is present.
if ngpleast_burnfile.endswith('.crdownload'):
    print('YOU FUCKED UP')
    print(ngpleast_dpath)
    print(ngpleast_burnfile_k)
    b=ngpleast_burnfile.split('.')[0:-1]
    b=b[0]+'.'+b[1]
    new_path=ngpleast_dpath+'/'+ b
    print(b)
    print(new_path)
    os.rename(ngpleast_burnfile,newpath)
else:
    pass
'''


pd.set_option('display.max_columns',1000)
pd.set_option('display.max_rows',1000)

file=os.listdir(ngpleast_dpath)[0] # selects file from burner directory
#print(ngpleast_burnfile_k) # prints name of the file to ensure .crdownload is not at the end
ngpleast_dfs = pd.read_excel(ngpleast_dpath+'/'+file,header=3) # ingests excel file into DataFrame
ngpleast_tsq = int(ngpleast_dfs[ngpleast_dfs['Loc']==46622]['Total Scheduled Quantity'].values[0]) # specific TSQ
#print(ngpleast_dfs) # prints tsq

print(ngpleast_tsq)

os.remove(ngpleast_burnfile) # removes burner excel file
os.rmdir(ngpleast_dpath) # removes burner directory




# Below is the Excel Automation

# Below is the counter code
joe_counter_file = open("joecounter.txt","r+") # Opens counter file
joecount = str(int(joe_counter_file.read())+1) # Reads file and takes the count, changes to integer, adds 1 and changes back to string
print(joecount)
joecounter = open('joecounter.txt', 'r+') # read state
joecounter.truncate(0) # Deletes present count
joecounterfile = open("joecounter.txt","w") # write state
joecounterfile.write(joecount) # writes updated number to .txt file
joecounterfile.close() # closes file

# Below is the DataFrame library structure

values = { 'LOC': ['KMLP','Transco Gulf Trace','Trunkline to Creole Trail','Transco to Creole Trail','TETCO to Creole Trail','Creole Trail SPLIQ','TGP to Corpus Christi','Transco to Corpus Christi','NGPL to Corpus Christi','KM Tejas to Corpus Christi','EPD Texas to Corpus Christi','Corpus Christi Liquefaction','Texas Eastern to CIP','Texas Eastern','Tenessee Gas Pipeline to CIP','Deliveries to Cameron Terminal','CGT Cameron Access','Elba','Transco Pleasant Valley','Columbia Loudoun','Cove Point Deliveries','Gulf South','TETCO','NGPL East'],
           'TSQ': [spkmlp_tsq,sbtgt_lighthouse,ct_trunkline_ctpl,ct_transco_ctpl,ct_tetco_ctpl,ct_spliq_ctpl,cc_tgp_ccpl,cc_transco_ccpl,cc_ngpl_ccpl,cc_kmtejas_ccpl,cc_epdtexas_ccpl,cc_ccliq,ci_texeastcip,ci_texeastern,ci_tenngaspipe,ci_termdeliv,cgt_tsq,elba_tsq,cp_trans_pleas_tsq,cp_col_lou_tsq,cp_cove_poi_tsq,gulf_south_tsq,last_tetco_tsq,ngpleast_tsq]}

# Above dictionary has zipped the values of LOC to TSQ into a structure readable by Pandas to create a DataFrame with neatly organized loc and tsq values

dframe = pd.DataFrame(values,columns = ['LOC','TSQ']) # Creates a DataFrame
print(dframe) # Prints DataFrame
dframe.to_excel('FeedGasBurner.xlsx') # Creates a burner Excel sheet with dataframe appended


burner = openpyxl.load_workbook("/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/FeedGasBurner.xlsx") # assigns this file to variable 'burner'
burn = burner['Sheet1'] # Assigns the proper sheet from burner workbook
feedgas = openpyxl.load_workbook('/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/Copy of Feed gas database_v30_FreeportTETCO_Raj - c1.xlsx') # Main file assigned to variable 'feedgas'
fg = feedgas['Volumes (input sheet)'] # Main sheet assigned to variable 'fg'

# Below are the cells from the burner DataFrame that contain the TSQ values. Values are converted to integers and formatted properly.
kmlp_sabine = int(burn.cell(column=3,row=2).value.replace(',',''))
transco_gulf_trace = int(burn.cell(column=3,row=3).value.replace(',',''))
trunk_to_creole = int(burn.cell(column=3,row=4).value.replace(',',''))
transco_to_creole = int(burn.cell(column=3,row=5).value.replace(',',''))
tetco_to_creole = int(burn.cell(column=3,row=6).value.replace(',',''))
creole_spliq = int(burn.cell(column=3,row=7).value.replace(',',''))
tgp_to_corpus = int(burn.cell(column=3,row=8).value.replace(',',''))
transco_to_corpus = int(burn.cell(column=3,row=9).value.replace(',',''))
ngpl_to_corpus = int(burn.cell(column=3,row=10).value.replace(',',''))
kmtejas_to_corpus = int(burn.cell(column=3,row=11).value.replace(',',''))
epdtexas_to_corpus = int(burn.cell(column=3,row=12).value.replace(',',''))
corpus_ccliq = int(burn.cell(column=3,row=13).value.replace(',',''))
texaseastern_to_cip = int(burn.cell(column=3,row=14).value.replace(',',''))
texaseastern = int(burn.cell(column=3,row=15).value.replace(',',''))
tgp_to_cip = int(burn.cell(column=3,row=16).value.replace(',',''))
cameron_deliv = int(burn.cell(column=3,row=17).value.replace(',',''))
cgt_access = burn.cell(column=3,row=18).value # value coming already an integer
elba_km = int(burn.cell(column=3,row=19).value.replace(',',''))
cp_transco_pleasant = int(burn.cell(column=3,row=20).value) # no comma
cp_columbia_loudoun = int(burn.cell(column=3,row=21).value) # no comma
cp_cove_deliveries = int(burn.cell(column=3,row=22).value) # ^
gulf_south = int(burn.cell(column=3,row=23).value)
tetco_tsq = int(burn.cell(column=3,row=24).value)
ngpl_east = int(burn.cell(column=3,row=25).value)

'''
# Ignore, Below are the cells which should be changed everyday to the next row. Pre automation work.
#fg['C1647'] = kmlp_sabine
#fg['D1647'] = transco_gulf_trace
#fg['G1647'] = trunk_to_creole
#fg['H1647'] = transco_to_creole
#fg['I1647'] = tetco_to_creole
#fg['J1647'] = creole_spliq
#fg['P1647'] = tgp_to_corpus
#fg['Q1647'] = transco_to_corpus
#fg['R1647'] = ngpl_to_corpus
#fg['S1647'] = kmtejas_to_corpus
#fg['T1647'] = epdtexas_to_corpus
#fg['U1647'] = corpus_ccliq
#fg['X1647'] = texaseastern_to_cip
#fg['Y1647'] = texaseastern
#fg['Z1647'] = tgp_to_cip
#fg['AC1647'] = cameron_deliv
#fg['AD1647'] = cgt_access
#fg['AO1647'] = elba_km
#fg['K1647'] = cp_transco_pleasant
#fg['L1647'] = cp_columbia_loudoun
#fg['M1647'] = cp_cove_deliveries
#fg['AW1647'] = gulf_south
#fg['AX1647'] = tetco_tsq
#fg['F1647'] = ngpl_east
'''
# Below are the cells that correspond to the Excel file. 
fg['C' + joecount] = kmlp_sabine
fg['D' + joecount] = transco_gulf_trace
fg['G' + joecount] = trunk_to_creole
fg['H' + joecount] = transco_to_creole
fg['I' + joecount] = tetco_to_creole
fg['J' + joecount] = creole_spliq
fg['P' + joecount] = tgp_to_corpus
fg['Q' + joecount] = transco_to_corpus
fg['R' + joecount] = ngpl_to_corpus
fg['S' + joecount] = kmtejas_to_corpus
fg['T' + joecount] = epdtexas_to_corpus
fg['U' + joecount] = corpus_ccliq
fg['X' + joecount] = texaseastern_to_cip
fg['Y' + joecount] = texaseastern
fg['Z' + joecount] = tgp_to_cip
fg['AC' + joecount] = cameron_deliv
fg['AD' + joecount] = cgt_access
fg['AO' + joecount] = elba_km
fg['K' + joecount] = cp_transco_pleasant
fg['L' + joecount] = cp_columbia_loudoun
fg['M' + joecount] = cp_cove_deliveries
fg['AW' + joecount] = gulf_south
fg['AX' + joecount] = tetco_tsq
fg['F' + joecount] = ngpl_east

feedgas.save(filename='/Users/rajgarkhedkar/Desktop/Enkon/2020/JoeAutomation/Copy of Feed gas database_v30_FreeportTETCO_Raj - c1.xlsx') # saves updated file
os.remove('FeedGasBurner.xlsx') # removes burner file from computer

print("Daily Scraping Complete!") # success!
