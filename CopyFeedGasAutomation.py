from bs4 import BeautifulSoup as soup
import pandas as pd
import openpyxl
import requests
#import urllib.request
#from pprint import pprint
import sys
#import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
#import time
#import ssl
#from lxml import etree
import os



# Below code is for Sabine Pass - Kinder Morgan KMLP
#print("Below are volumes for KMLP - Sabine Pass (Kinder Morgan)")
spkmlp_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Path must be specified depending on local machine
spkmlp_driver.get('https://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code=KMLP') # get request for Selenium to download the page in a proxy browser
spkmlp_retrieve_xpath = '//*[@id="WebSplitter1_tmpl1_ContentPlaceHolder1_HeaderBTN1_btnRetrieve"]' # ignore - xpath for retrieve button - debugging
spkmlp_retrieve_click = spkmlp_driver.find_element_by_name('ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnRetrieve').click() # automatically initiates a click on retrieve button to download content
spkmlp_tsq = spkmlp_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[4]/div[1]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[16]/td/div/table/tbody/tr/td[8]").get_property("innerText") # retrieves the TSQ values for KMLP Liquefaction Deliveries
spkmlp_driver.close() # closes opened proxy page

'''
# BautifulSoup and Pandas method, ignore
#spkmlp_content = spkmlp_driver.page_source # Gathers html content from retrieved informational posting
#spkmlp_html = soup(spkmlp_content, 'lxml') # begins a parser that can iterate through html content
#print(sp_html.prettify()) # Initial print to verify all html content is loaded

#spkmlp_loc = spkmlp_html.select('td.igb5ea3b') # Specific cell 'td' and class 'igb5ea3b' common tags for all loc cells
#spkmlp_tsq = spkmlp_html.select('td.igb053bb47') # Specific cell 'td' and class 'igb053bb47' common tags for all tsq cells

#spkmlp_data = {} # Empty dictionary to store the loc and tsq values
#for spkmlp_location, spkmlp_quant in zip(spkmlp_loc, spkmlp_tsq): # Merges the values together, ties loc with tsq
#    spkmlp_data[spkmlp_location.text] = spkmlp_quant.text # assigns a 1:1 loc to tsq relationship and appends to the dictionary
#pprint(spkmlp_data) # Pretty prints the selected loc and tsq data zipped.

# Below is the Excel automation using pandas

#spkmlp_df = pd.DataFrame(data=spkmlp_data, index=[0]) # Creates an empty dataframe and assigns the data values in a readable format for Excel
#spkmlp_df = (spkmlp_df.T)  # Transposes index and column
#print(spkmlp_df) # Prints output so user can ensure data validity
#df.to_excel('FeedGasScrapebook.xlsx') # Creates an excel file and appends all data into a column for loc and column for tsq. Unhash to use.
'''


# Below is script for Sabine Pass - TGT Lighthouse Road M4662 MP 13
#print("Below are volumes for TGT Lighthouse Road - Sabine Pass")
sbtgt_url = "http://www.1line.williams.com/ebbCode/OACQueryRequest.jsp?BUID=80&type=OAC" # URL
sbtgt_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
sbtgt_driver.get(sbtgt_url) # Driver proxy request
sbtgt_driver.find_element_by_css_selector("body > form > p > input[type=submit]").click() # retrieve button click
WebDriverWait(sbtgt_driver, 30).until(EC.number_of_windows_to_be(2)) # Driver wait request to prevent a timeout error
sbtgt_report = sbtgt_driver.window_handles[1] # Prioritizes the result window to initiate scrape
sbtgt_driver.switch_to_window(sbtgt_report) # Switches to window to enable scrape
#sbtgt_driver.switch_to_window("OACreport")
sbtgt_lighthouse = sbtgt_driver.find_element_by_xpath("/html/body/form/table/tbody/tr[2]/td/div/table[3]/tbody/tr[223]/td[9]").get_property("textContent") # Selects the TSQ and retrieves TSQ text
#print(sbtgt_lighthouse)
sbtgt_driver.close() # Closes result page
sbtgt_driver.quit() # Terminates proxy browser




'''
# Below is script for Sabine pass - NGPL East
print("Below are volumes for NGPL East - Sabine Pass (Kinder Morgan)")
spngpl_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Path must be specified depending on local machine
spngpl_driver.get('https://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code=NGPL') # get request for Selenium to download the page in a proxy browser
spngpl_retrieve_xpath = '//*[@id="WebSplitter1_tmpl1_ContentPlaceHolder1_HeaderBTN1_btnRetrieve"]' # ignore - xpath for retrieve button - debugging
#spngpl_retrieve_click = spngpl_driver.find_element_by_name('ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnRetrieve').click() # automatically initiates a click on retrieve button to download content
spngpl_download_click = spngpl_driver.find_element_by_name('ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnDownload').click() # automatically initiates a click on retrieve button to download content

spngpl_driver.close()
'''
'''
# Selenium for NGPL - full of errors, ignore.
#spngpl_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Path must be specified depending on local machine
#spngpl_driver.get('https://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code=NGPL') # get request for Selenium to download the page in a proxy browser
#spngpl_retrieve_xpath = '//*[@id="WebSplitter1_tmpl1_ContentPlaceHolder1_HeaderBTN1_btnRetrieve"]' # ignore - xpath for retrieve button - debugging
#spngpl_retrieve_click = spngpl_driver.find_element_by_name('ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnRetrieve').click() # automatically initiates a click on retrieve button to download content

# Below is incomplete automation. Kinder Morgan site is too unstable for full automation, fails most of the time. Excel download is much easier.
#spngpl_content = spngpl_driver.page_source # Gathers html content from retrieved informational posting
#spngpl_html = soup(spngpl_content, 'lxml') # begins a parser that can iterate through html content
#spngpl_driver.close() # closes opened proxy page
#print(spngpl_html.prettify()) # Initial print to verify all html content is loaded

#spngpl_loc = spngpl_html.select('td.igb5ea3b') # Specific cell 'td' and class 'igb5ea3b' common tags for all loc cells
#spngpl_tsq = spngpl_html.select('td.igb053bb47') # Specific cell 'td' and class 'igb053bb47' common tags for all tsq cells

#spngpl_data = {} # Empty dictionary to store the loc and tsq values
#for spngpl_location, spngpl_quant in zip(spngpl_loc, spngpl_tsq): # Merges the values together, ties loc with tsq
#    spngpl_data[spngpl_location.text] = spngpl_quant.text # assigns a 1:1 loc to tsq relationship and appends to the dictionary
#pprint(spngpl_data) # Pretty prints the selected loc and tsq data zipped.

# Below is the Excel automation using pandas

#spngpl_df = pd.DataFrame(data=spngpl_data, index=[0]) # Creates an empty dataframe and assigns the data values in a readable format for Excel
#spngpl_df = (spngpl_df.T)  # Transposes index and column
#print(spngpl_df) # Prints output so user can ensure data validity
#df.to_excel('FeedGasScrapebook.xlsx') # Creates an excel file and appends all data into a column for loc and column for tsq. Unhash to use
'''


# Below is the script for Sabine Pass - Creole Trail Cheniere LNG Connection
#print("Below are volumes for Creole Trail - Sabine Pass (LNG Connection)")
ct_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # filepath for driver
ct_driver.get('https://lngconnection.cheniere.com/#/ctpl') # Opens proxy browser and fetches URL
ct_elems = WebDriverWait(ct_driver, 30).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="capacityGrid"]/table/tbody/tr[1]/td[4]'))) # Wait request to prevent timeout error
ct_driver.find_element_by_xpath("/html/body/app/main/side-menu/nav/ul/li[2]/a").click() # Clicks on Capacity
ct_driver.find_element_by_xpath("/html/body/app/main/side-menu/nav/ul/li[2]/ul/li[1]/a").click() # Clicks on Operationally Available
ct_table_elems = WebDriverWait(ct_driver, 30).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="gridOperationallyAvailable"]/table/tbody/tr[1]/td[5]'))) # Wait request to finish loading content
#ct_html = ct_driver.page_source

ct_trunkline_ctpl = ct_driver.find_element_by_xpath('//*[@id="gridOperationallyAvailable"]/table/tbody/tr[3]/td[14]').get_property('innerText') # TSQ
ct_transco_ctpl = ct_driver.find_element_by_xpath('//*[@id="gridOperationallyAvailable"]/table/tbody/tr[2]/td[14]').get_property('innerText') # TSQ
ct_tetco_ctpl = ct_driver.find_element_by_xpath('//*[@id="gridOperationallyAvailable"]/table/tbody/tr[1]/td[14]').get_property('innerText') # TSQ
ct_spliq_ctpl = ct_driver.find_element_by_xpath('//*[@id="gridOperationallyAvailable"]/table/tbody/tr[4]/td[14]').get_property('innerText') # TSQ
ct_driver.close()

'''
# BeautifulSoup method for debugging, ignore
#ct_content = soup(ct_html, 'html.parser')
#ct_driver.close()
#ct_table = ct_content.find('table', class_='k-selectable')
#print(ct_content.prettify())
#print(ct_table.prettify())
#print(ct_trunkline_ctpl)
#print(ct_transco_ctpl)
#print(ct_tetco_ctpl)
#print(ct_spliq_ctpl)
'''


# Below is the script for Cove Point LNG
#print("Below are volumes for Cove Point LNG")
#cp_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver')
#cp_driver.get('https://dekaflow.dominionenergy.com/servlet/InfoPostServlet?region=null&company=cpt&method=headers&category=Capacity&subcategory=Operationally+Available')
#cp_driver.find_element_by_name("").click() # Initiates download




# Below is the script for Corpus Christi - LNG Connection
#print("Below are volumes for Corpus Christi - LNG Connection")
cc_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
cc_driver.get('https://lngconnection.cheniere.com/#/ccpl') # Opens proxy page with the URL
cc_elems = WebDriverWait(cc_driver, 10).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="capacityGrid"]/table/tbody/tr[1]/td[4]'))) # Wait request for content to load
#cc_html = cc_driver.page_source


cc_tgp_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[2]/td[4]').get_property('innerText') # TSQ
cc_transco_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[3]/td[4]').get_property('innerText') # TSQ
cc_ngpl_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[4]/td[4]').get_property('innerText') # TSQ
cc_kmtejas_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[6]/td[4]').get_property('innerText') # TSQ
cc_epdtexas_ccpl = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[5]/td[4]').get_property('innerText') # TSQ
cc_ccliq = cc_driver.find_element_by_xpath('//*[@id="capacityGrid"]/table/tbody/tr[7]/td[4]').get_property('innerText') # TSQ
cc_driver.close()

'''
# Ignore below code - debugging
#cc_content = soup(cc_html, 'html.parser')
#cc_driver.close()
#cc_table = cc_content.find('table', class_='k-selectable')
#print(cc_content.prettify())
#print(cc_table.prettify())
#print(cc_tgp_ccpl, ":TGP Sinton R")
#print(cc_transco_ccpl, ":Transco San Pat R")
#print(cc_ngpl_ccpl, ":NGPL Sinton R")
#print(cc_kmtejas_ccpl, ":KM Tejas Sinton R")
#print(cc_epdtexas_ccpl, ":EPROD San Pat R")
#print(cc_ccliq, ":Corpus Christi CCLIQ ")

#cc_loc = cc_content.select('td.igb5ea3b') # Specific cell 'td' and class 'igb5ea3b' common tags for all loc cells
#cc_tsq = cc_content.select('td.igb053bb47') # Specific cell 'td' and class 'igb053bb47' common tags for all tsq cells

#cc_data = {} # Empty dictionary to store the loc and tsq values
#for cc_location, cc_quant in zip(cc_loc, cc_tsq): # Merges the values together, ties loc with tsq
    #cc_data[cc_location.text] = cc_quant.text # assigns a 1:1 loc to tsq relationship and appends to the dictionary
#pprint(cc_data) # Pretty prints the selected loc and tsq data zipped.
'''



# Below is the script for Cameron LNG

ci_url = "http://www.gasnom.com/ip/cameron/oauc.cfm?type=1" # URL for Gasnom
ci_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for driver
ci_driver.get(ci_url) # Opens proxy page for the URL
ci_texeastcip = ci_driver.find_element_by_xpath("/html/body/div[1]/div[3]/table[2]/tbody/tr[13]/td[9]").get_property("innerText") # TSQ
ci_texeastern = ci_driver.find_element_by_xpath("/html/body/div[1]/div[3]/table[2]/tbody/tr[15]/td[9]").get_property("innerText") # TSQ
ci_tenngaspipe = ci_driver.find_element_by_xpath("/html/body/div[1]/div[3]/table[2]/tbody/tr[10]/td[9]").get_property("innerText") # TSQ
ci_termdeliv = ci_driver.find_element_by_xpath("/html/body/div[1]/div[3]/table[2]/tbody/tr[3]/td[9]").get_property("innerText") # TSQ
ci_driver.close() # Closes proxy page

'''
# Pandas method, ignore
#print("Below are volumes for Cameron LNG")
#ci_url = 'http://www.gasnom.com/ip/cameron/oauc.cfm?type=1'
#pd.set_option('display.max_rows', 500)
#pd.set_option('display.max_columns', 500)
#pd.set_option('display.width', 1000)
#ci_dfs = pd.read_html(ci_url)[2]
#print(ci_dfs)
'''

# Below is the volume for CGT Cameron Access

cgt_url = "http://www.columbiapipeinfo.com/infopost/ReportViewer.aspx?/InfoPost/OperationallyAvailableCapacity&pAssetNbr=51" # URL
cgt_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
cgt_driver.get(cgt_url) # Opens proxy page with URL
cgt_elems = WebDriverWait(cgt_driver, 10).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr[4]/td[3]/div/div[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[8]/td[3]/table/tbody/tr[3]/td[1]/div/div'))) # Waits for content to finish loading to prevent timeout error
cgt_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[1]/td[2]/div/select").click() # Clicks on asset
cgt_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[1]/td[2]/div/select/option[1]").click() # Selects CGT Asset
try: # Prevent stale DOM error
    cgt_reportwait = WebDriverWait(cgt_driver, 5).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[1]/td[2]/div/select'))) # Waits for report button to be visible for iterator
except: # Prevent stale DOM error
    pass # Prevent stale DOM error
cgt_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[3]/table/tbody/tr/td/input").submit() # Clicks on view report and submits form
cgt_findwait = WebDriverWait(cgt_driver, 20).until(EC.element_to_be_clickable((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[3]/td[2]/div/select'))) # Wait request for content to load to prevent timeout error
cgt_findwait.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[3]/td[2]/div/select").click() # Clicks on location
cgt_findwait.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[1]/table/tbody/tr[3]/td[2]/div/select/option[20]").click() # Clicks on Cameron Pipeline (loc: 4246)
cgt_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr[1]/td/div/div/table/tbody/tr/td[3]/table/tbody/tr/td/input").click() # Clicks on view report
cgt_tsq = WebDriverWait(cgt_driver, 10).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr[4]/td[3]/div/div[1]/div/table/tbody/tr/td/table/tbody/tr[2]/td/table/tbody/tr/td/table/tbody/tr[8]/td[3]/table/tbody/tr[4]/td[6]/div/div/div'))).get_property("innerText") # Wait request to finish loading page and then captures TSQ value
cgt_driver.close() # Closes proxy page


# Below is the script for Elba - Elba Express Connector (ELC Chatham Deliveries) - ELC/EEC EEC TO ELC CHATHAM 660700 - Kinder Morgan
#print("Below are volumes for Elba Express Chatham Deliveries - Kinder Morgan")
elba_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Path must be specified depending on local machine
elba_driver.get('https://pipeline2.kindermorgan.com/Capacity/OpAvailPoint.aspx?code=EEC') # get request for Selenium to download the page in a proxy browser
elba_driver.find_element_by_name("ctl00$WebSplitter1$tmpl1$ContentPlaceHolder1$HeaderBTN1$btnRetrieve").click() # Clicks on retrieve button
elba_elems = WebDriverWait(elba_driver, 10).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[4]/div[1]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[3]/td/div/table/tbody/tr/td[8]'))) # Wait request to finish loading content
elba_tsq = elba_driver.find_element_by_xpath("/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[4]/div[1]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[3]/td/div/table/tbody/tr/td[8]").get_property("innerText") # captures TSQ of point specified
#print(elba_tsq)
elba_driver.close() # Closes proxy page

# Print commands for all variables with TSQ values stored
print("Below are volumes for Sabine Pass KMLP")
print(spkmlp_tsq)
print("Below are volumes for TGT Lighthouse")
print(sbtgt_lighthouse)
print("Below are volumes for Sabine Pass - Creole Trail")
print(ct_trunkline_ctpl)
print(ct_transco_ctpl)
print(ct_tetco_ctpl)
print(ct_spliq_ctpl)
print("Below are volumes for Corpus Christi")
print(cc_tgp_ccpl)
print(cc_transco_ccpl)
print(cc_ngpl_ccpl)
print(cc_kmtejas_ccpl)
print(cc_epdtexas_ccpl)
print(cc_ccliq)
print("Below are volumes for Cameron Interstate")
print(ci_texeastcip)
print(ci_texeastern)
print(ci_tenngaspipe)
print(ci_termdeliv)
print("Below is the volume for CGT Cameron Access")
print(cgt_tsq)
print("Below are volumes for Elba - Chatham Deliveries")
print(elba_tsq)

# Below is the DataFrame library structure

values = { 'LOC': ['KMLP','Transco Gulf Trace','Trunkline to Creole Trail','Transco to Creole Trail','TETCO to Creole Trail','Creole Trail SPLIQ','TGP to Corpus Christi','Transco to Corpus Christi','NGPL to Corpus Christi','KM Tejas to Corpus Christi','EPD Texas to Corpus Christi','Corpus Christi Liquefaction','Texas Eastern to CIP','Texas Eastern','Tenessee Gas Pipeline to CIP','Deliveries to Cameron Terminal','CGT Cameron Access','Elba'],
           'TSQ': [spkmlp_tsq,sbtgt_lighthouse,ct_trunkline_ctpl,ct_transco_ctpl,ct_tetco_ctpl,ct_spliq_ctpl,cc_tgp_ccpl,cc_transco_ccpl,cc_ngpl_ccpl,cc_kmtejas_ccpl,cc_epdtexas_ccpl,cc_ccliq,ci_texeastcip,ci_texeastern,ci_tenngaspipe,ci_termdeliv,cgt_tsq,elba_tsq]}
# Above dictionary has zipped the values of LOC to TSQ into a structure readable by Pandas to create a DataFrame with neatly organized loc and tsq values
dframe = pd.DataFrame(values,columns = ['LOC','TSQ']) # Creates a DataFrame
print(dframe) # Prints DataFrame
dframe.to_excel('FeedGasBurner.xlsx') # Creates a burner Excel sheet with dataframe appended


# Below is the Excel Automation
burner = openpyxl.load_workbook("/Users/rajgarkhedkar/Desktop/Enkon/2020/FeedGasBurner.xlsx") # assigns this file to variable 'burner'
burn = burner['Sheet1'] # Assigns the proper sheet from burner workbook
feedgas = openpyxl.load_workbook('/Users/rajgarkhedkar/Desktop/Enkon/2020/Copy of Feed gas database_v30_FreeportTETCO_Raj - c1.xlsx') # Main file assigned to variable 'feedgas'
fg = feedgas['Volumes (input sheet)'] # Main sheet assigned to variable 'fg'

# Below are the cells from the burner DataFrame that contain the TSQ values.
kmlp_sabine = burn.cell(column=3,row=2).value
transco_gulf_trace = burn.cell(column=3,row=3).value
trunk_to_creole = burn.cell(column=3,row=4).value
transco_to_creole = burn.cell(column=3,row=5).value
tetco_to_creole = burn.cell(column=3,row=6).value
creole_spliq = burn.cell(column=3,row=7).value
tgp_to_corpus = burn.cell(column=3,row=8).value
transco_to_corpus = burn.cell(column=3,row=9).value
ngpl_to_corpus = burn.cell(column=3,row=10).value
kmtejas_to_corpus = burn.cell(column=3,row=11).value
epdtexas_to_corpus = burn.cell(column=3,row=12).value
corpus_ccliq = burn.cell(column=3,row=13).value
texaseastern_to_cip = burn.cell(column=3,row=14).value
texaseastern = burn.cell(column=3,row=15).value
tgp_to_cip = burn.cell(column=3,row=16).value
cameron_deliv = burn.cell(column=3,row=17).value
cgt_access = burn.cell(column=3,row=18).value
elba_km = burn.cell(column=3,row=19).value

# Below are the cells which should be changed everyday to the next row.
fg['C1647'] = kmlp_sabine
fg['D1647'] = transco_gulf_trace
fg['G1647'] = trunk_to_creole
fg['H1647'] = transco_to_creole
fg['I1647'] = tetco_to_creole
fg['J1647'] = creole_spliq
fg['P1647'] = tgp_to_corpus
fg['Q1647'] = transco_to_corpus
fg['R1647'] = ngpl_to_corpus
fg['S1647'] = kmtejas_to_corpus
fg['T1647'] = epdtexas_to_corpus
fg['U1647'] = corpus_ccliq
fg['X1647'] = texaseastern_to_cip
fg['Y1647'] = texaseastern
fg['Z1647'] = tgp_to_cip
fg['AC1647'] = cameron_deliv
fg['AD1647'] = cgt_access
fg['AO1647'] = elba_km

feedgas.save(filename='/Users/rajgarkhedkar/Desktop/Enkon/2020/Copy of Feed gas database_v30_FreeportTETCO_Raj - c1.xlsx')
os.remove('FeedGasBurner.xlsx')

print("Daily Scraping Complete!")
