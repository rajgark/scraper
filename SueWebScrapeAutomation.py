from bs4 import BeautifulSoup as soup
import pandas as pd
import openpyxl
import requests
#import urllib.request
#from pprint import pprint
import sys
#import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
#import time
#import ssl
#from lxml import etree
import os

'''

Northern Natural Gas
https://www.northernnaturalgas.com/infopostings/Capacity/Pages/OperationallyAvailable.aspx
LOC: 1022 Brownfield North Group D
Flow Desc: Permian-TX to Mid-Con

NGPL - segment capacity
https://pipeline2.kindermorgan.com/Capacity/OpAvailSegment.aspx?code=NGPL
LOC: 8 North of Sta. 167

Transwestern Pipeline
https://twtransfer.energytransfer.com/ipost/TW/capacity/operationally-available
LOC: M2 of WT-1 System

El Paso - segment capacity
https://pipeline2.kindermorgan.com/Capacity/OpAvailSegment.aspx?code=EPNG
LOC: Caprock N, Perm N, CornHPW, CornLPW, Corntrn, CornL2K
       640       940     1600     1600     1600     1600

webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
WebDriverWait(ct_driver, 30).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="capacityGrid"]/table/tbody/tr[1]/td[4]'))) # template for WebDriverWait, replace as needed

'''


# Below is the script for Northern Natural Gas - Brownfield North Group D
nngas_url = 'https://www.northernnaturalgas.com/infopostings/Capacity/Pages/OperationallyAvailable.aspx'
nngas_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
nngas_driver.get(nngas_url)
nngas_driver.find_element_by_name('ctl00$ctl45$g_a9dc7b41_6f27_4c06_88b9_657585c54828$ctl00$ctl03').click() # Initiates click on 'Type of Search'
nngas_driver.find_element_by_xpath("//option[@value='allgroups']").click() # Selects on "All Group Locations"
nngas_driver.find_element_by_name('ctl00$ctl45$g_a9dc7b41_6f27_4c06_88b9_657585c54828$ctl00$ctl02').click() # Initiates click on 'Cycle'
nngas_driver.find_element_by_xpath("//option[@value='10']").click() # Selects 'Timely'
nngas_driver.find_element_by_name('ctl00$ctl45$g_a9dc7b41_6f27_4c06_88b9_657585c54828$ctl00$ctl08').click() # Retrieves information
# NOT full xpath used below
nngas_designcap = nngas_driver.find_element_by_xpath('//*[@id="ctl00_ctl45_g_a9dc7b41_6f27_4c06_88b9_657585c54828_ctl00_dg_OAC_NAESB30_ctl00__80"]/td[8]').get_property('innerText')
nngas_opcap = nngas_driver.find_element_by_xpath('//*[@id="ctl00_ctl45_g_a9dc7b41_6f27_4c06_88b9_657585c54828_ctl00_dg_OAC_NAESB30_ctl00__80"]/td[9]').get_property('innerText') # Captures Operating Capacity
nngas_tsq = nngas_driver.find_element_by_xpath('//*[@id="ctl00_ctl45_g_a9dc7b41_6f27_4c06_88b9_657585c54828_ctl00_dg_OAC_NAESB30_ctl00__80"]/td[10]').get_property('innerText') # Captures TSQ
nngas_opavail = nngas_driver.find_element_by_xpath('//*[@id="ctl00_ctl45_g_a9dc7b41_6f27_4c06_88b9_657585c54828_ctl00_dg_OAC_NAESB30_ctl00__80"]/td[11]').get_property('innerText') # Captures Operationally Available Capacity

print("Brownfield North Group Design Capacity: ",nngas_designcap)
print("Brownfield North Group Operating Capacity: ",nngas_opcap)
print("Brownfield North Group Total Scheduled Capacity: ",nngas_tsq)
print("Brownfield North Group Operationally Available Capacity: ",nngas_opavail)
#nngas_html = nngas_driver.page_source # Ignore, captures HTML for debugging
#print(nngas_html) # Ignore, prints HTML for debugging
nngas_driver.close() # Closes driver
print("-----") # Separation dashes



# Below is the script for Kinder Morgan NGPL - North of Sta. 167 LOC Segment 8
ngpl_url = 'https://pipeline2.kindermorgan.com/Capacity/OpAvailSegment.aspx?code=NGPL'
ngpl_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
ngpl_driver.get(ngpl_url)
ngpl_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[1]/table[1]/tbody/tr/td[2]/table/tbody/tr/td/input[1]').click() # Clicks on 'Retrieve'
WebDriverWait(ngpl_driver, 30).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[1]/td/div/table/tbody/tr/td[3]'))) # Wait request for content to load
# FULL xpath copied below
ngpl_designcap = ngpl_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[7]/td/div/table/tbody/tr/td[5]').get_property('innerText')
ngpl_designcap = int(ngpl_designcap.replace(',',''))
ngpl_opcap = ngpl_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[7]/td/div/table/tbody/tr/td[6]').get_property('innerText') # Stores operationally available capacity
ngpl_opcap = int(ngpl_opcap.replace(',',''))
ngpl_tsq = ngpl_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[7]/td/div/table/tbody/tr/td[7]').get_property('innerText') # Stores tsq
ngpl_tsq = int(ngpl_tsq.replace(',',''))
ngpl_opavailcap = ngpl_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[7]/td/div/table/tbody/tr/td[8]').get_property('innerText') # Stores operationally available capacity
ngpl_opavailcap = int(ngpl_opavailcap.replace(',',''))

ngpl_driver.close()

print("North of Sta. 167 Design Capacity: ",ngpl_designcap)
print("North of Sta. 167 Operating Capacity: ",ngpl_opcap)
print("North of Sta. 167 Total Scheduled Quantity: ",ngpl_tsq)
print("North of Sta. 167 Operationally Available Capacity",ngpl_opavailcap)
print("-----")



# Below is the script for Transwestern Pipeline - M2 of WT-1 System
tpc_url = 'https://twtransfer.energytransfer.com/ipost/TW/main/index'
tpc_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
tpc_driver.get(tpc_url)
WebDriverWait(tpc_driver, 30).until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/div/article/section[2]/div/div[2]/div/table/tbody[6]/tr[2]/td[4]'))) # template for WebDriverWait, replace as needed

# FULL xpath copied below
wt1_opcap = tpc_driver.find_element_by_xpath('/html/body/div/div/article/section[2]/div/div[2]/div/table/tbody[6]/tr[2]/td[4]').get_property('innerText')
wt1_tsq = tpc_driver.find_element_by_xpath('/html/body/div/div/article/section[2]/div/div[2]/div/table/tbody[6]/tr[2]/td[5]').get_property('innerText')
wt1_opavailcap = tpc_driver.find_element_by_xpath('/html/body/div/div/article/section[2]/div/div[2]/div/table/tbody[6]/tr[2]/td[6]').get_property('innerText')
tpc_driver.close()

print("WT-1 System M2 Operating Capacity: ",wt1_opcap)
print("WT-1 System M2 Total Scheduled Quantity: ",wt1_tsq)
print("WT-1 System M2 Operationally Available Capacity: ",wt1_opavailcap)
print("-----")



# Below is the script for Kinder Morgan El Paso - Segments: Caprock N (640), Perm N (940), CornHPW (1600), CornLPW (1600), Corntrn (1600), CornL2K (1600)
elpaso_url = 'https://pipeline2.kindermorgan.com/Capacity/OpAvailSegment.aspx?code=EPNG'
elpaso_driver = webdriver.Chrome('/Users/rajgarkhedkar/Desktop/chromedriver') # Filepath for chromedriver
elpaso_driver.get(elpaso_url)
elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[1]/table[1]/tbody/tr/td[2]/table/tbody/tr/td/input[1]').click() # Clicks on 'Retrieve'
WebDriverWait(elpaso_driver, 30).until(EC.visibility_of_element_located((By.XPATH,'/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[1]/td/div/table/tbody/tr/td[3]'))) # Wait request for content to load

# FULL xpath copied below
caprock_designcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[16]/td/div/table/tbody/tr/td[5]').get_property('innerText')
caprock_opcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[16]/td/div/table/tbody/tr/td[6]').get_property('innerText') # Stores operationally available capacity
caprock_tsq = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[16]/td/div/table/tbody/tr/td[7]').get_property('innerText') # Stores tsq
caprock_opavailcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[16]/td/div/table/tbody/tr/td[8]').get_property('innerText') # Stores operationally available capacity

permn_designcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[65]/td/div/table/tbody/tr/td[5]').get_property('innerText')
permn_opcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[65]/td/div/table/tbody/tr/td[6]').get_property('innerText') # Stores operationally available capacity
permn_tsq = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[65]/td/div/table/tbody/tr/td[7]').get_property('innerText') # Stores tsq
permn_opavailcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[65]/td/div/table/tbody/tr/td[8]').get_property('innerText') # Stores operationally available capacity

cornhpw_designcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[69]/td/div/table/tbody/tr/td[5]').get_property('innerText')
cornhpw_opcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[69]/td/div/table/tbody/tr/td[6]').get_property('innerText') # Stores operationally available capacity
cornhpw_tsq = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[69]/td/div/table/tbody/tr/td[7]').get_property('innerText') # Stores tsq
cornhpw_opavailcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[69]/td/div/table/tbody/tr/td[8]').get_property('innerText') # Stores operationally available capacity

cornlpw_designcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[70]/td/div/table/tbody/tr/td[5]').get_property('innerText')
cornlpw_opcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[70]/td/div/table/tbody/tr/td[6]').get_property('innerText') # Stores operationally available capacity
cornlpw_tsq = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[70]/td/div/table/tbody/tr/td[7]').get_property('innerText') # Stores tsq
cornlpw_opavailcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[70]/td/div/table/tbody/tr/td[8]').get_property('innerText') # Stores operationally available capacity

corntrn_designcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[71]/td/div/table/tbody/tr/td[5]').get_property('innerText')
corntrn_opcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[71]/td/div/table/tbody/tr/td[6]').get_property('innerText') # Stores operationally available capacity
corntrn_tsq = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[71]/td/div/table/tbody/tr/td[7]').get_property('innerText') # Stores tsq
corntrn_opavailcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[71]/td/div/table/tbody/tr/td[8]').get_property('innerText') # Stores operationally available capacity

cornl2k_designcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[72]/td/div/table/tbody/tr/td[5]').get_property('innerText')
cornl2k_opcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[72]/td/div/table/tbody/tr/td[6]').get_property('innerText') # Stores operationally available capacity
cornl2k_tsq = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[72]/td/div/table/tbody/tr/td[7]').get_property('innerText') # Stores tsq
cornl2k_opavailcap = elpaso_driver.find_element_by_xpath('/html/body/form/div[3]/div/div/table/tbody/tr/td[3]/div/div/div[3]/table/tbody/tr[2]/td[1]/table/tbody[2]/tr/td/div[2]/table/tbody/tr[72]/td/div/table/tbody/tr/td[8]').get_property('innerText') # Stores operationally available capacity

elpaso_driver.close()

print("Caprock N Design Capacity: ",caprock_designcap)
print("Caprock N Operating Capacity: ",caprock_opcap)
print("Caprock N Total Scheduled Quantity: ",caprock_tsq)
print("Caprock N Operationally Available Capacity: ",caprock_opavailcap)
print(" ")
print("Perm N Design Capacity: ",permn_designcap)
print("Perm N Operating Capacity: ",permn_opcap)
print("Perm N Total Scheduled Quantity: ",permn_tsq)
print("Perm N Operationally Available Capacity: ",permn_opavailcap)
print(" ")
print("CornHPW Design Capacity: ",cornhpw_designcap)
print("CornHPW Operating Capacity: ",cornhpw_opcap)
print("CornHPW Total Scheduled Capacity: ",cornhpw_tsq)
print("CornHPW Operationally Available Capacity: ",cornhpw_opavailcap)
print(" ")
print("CornLPW Design Capacity: ",cornlpw_designcap)
print("CornLPW Operating Capacity: ",cornlpw_opcap)
print("CornLPW Total Scheduled Quantity: ", cornlpw_tsq)
print("CornLPW Operationally Available Capacity: ",cornlpw_opavailcap)
print(" ")
print("Corntrn Design Capacity: ",corntrn_designcap)
print("Corntrn Operating Capacity: ",corntrn_opcap)
print("Corntrn Total Scheduled Quantity: ",corntrn_tsq)
print("Corntrn Operationally Available Capacity: ", corntrn_opavailcap)
print(" ")
print("CornL2K Design Capacity: ",cornl2k_designcap)
print("CornL2K Operating Capacity: ",cornl2k_opcap)
print("CornL2K Total Scheduled Capacity: ",cornl2k_tsq)
print("CornL2K Operationally Available Capacity: ",cornl2k_opavailcap)



# Below is the excel automation


# Northern Natural Gas Counter
sue_northern_counter = open("suenorthernnaturalcounter.txt","r+") # Opens counter file
suenngcounter = str(int(sue_northern_counter.read())+1) # Reads file and takes the count, changes to integer, adds 1 and changes back to string
print(suenngcounter)
suenortherncounter = open('suenorthernnaturalcounter.txt', 'r+') # read state
suenortherncounter.truncate(0) # Deletes present count
suenortherncounterfile = open("suenorthernnaturalcounter.txt","w") # write state
suenortherncounterfile.write(suenngcounter) # writes updated number to .txt file
suenortherncounterfile.close() # closes file

# NGPL Counter
ngpl_counter_file = open("suengplcounter.txt","r+") # Opens counter file
suengplcounter = str(int(ngpl_counter_file.read())+1) # Reads file and takes the count, changes to integer, adds 1 and changes back to string
print(suengplcounter)
ngpl_count = open('suengplcounter.txt', 'r+') # read state
ngpl_count.truncate(0) # Deletes present count
suengplcounterfile = open("suengplcounter.txt","w") # write state
suengplcounterfile.write(suengplcounter) # writes updated number to .txt file
suengplcounterfile.close() # closes file

# Transwestern Counter
sue_trans_counter_file = open("suetranswesterncounter.txt","r+") # Opens counter file
transcount = str(int(sue_trans_counter_file.read())+1) # Reads file and takes the count, changes to integer, adds 1 and changes back to string
print(transcount)
transcounter = open('suetranswesterncounter.txt', 'r+') # read state
transcounter.truncate(0) # Deletes present count
suetranscounterfile = open("suetranswesterncounter.txt","w") # write state
suetranscounterfile.write(transcount) # writes updated number to .txt file
suetranscounterfile.close() # closes file

# El Paso Counter
elpaso_counter_file = open("sueelpasocounter.txt","r+") # Opens counter file
elpasocount = str(int(elpaso_counter_file.read())+1) # Reads file and takes the count, changes to integer, adds 1 and changes back to string
print(elpasocount)
suecounter = open('sueelpasocounter.txt', 'r+') # read state
suecounter.truncate(0) # Deletes present count
elpasocounterfile = open("sueelpasocounter.txt","w") # write state
elpasocounterfile.write(elpasocount) # writes updated number to .txt file
elpasocounterfile.close() # closes file



# Below is a library structure
values = { 'LOC': ['Brownfield North Group D Design Capacity','Brownfield North Group D Operating Capacity','Brownfield North Group D Total Scheduled Quantity','Brownfield North Group D Operationally Available Capacity','North of Sta. 167 Design Capacity','North of Sta. 167 Operating Capacity','North of Sta. 167 Total Scheduled Quantity','North of Sta. 167 Operationally Available Capacity','WT-1 System M2 Operating Capacity','WT-1 System M2 Total Scheduled Quanitity','WT-1 System M2 Operationally Available Capacity','Caprock N Design Capacity','Caprock N Operating Capacity','Caprock N Total Scheduled Quantity','Caprock N Operationally Available Capacity','Perm N Design Capacity','Perm N Operating Capacity','Perm N Total Scheduled Quantity','Perm N Operationally Available Capacity','CornHPW Design Capacity','CornHPW Operating Capacity','CornHPW Total Scheduled Quantity','CornHPW Operationally Available Capacity','CornLPW Design Capacity','CornLPW Operating Capacity','CornLPW Total Scheduled Quantity','CornLPW Operationally Available Capacity','Corntrn Design Capacity','Corntrn Operating Capacity','Corntrn Total Scheduled Quantity','Corntrn Operationally Available Capacity','CornL2K Design Capacity','CornL2K Operating Capacity','CornL2K Total Scheduled Quantity','CornL2K Operationally Available Capacity'],
           'TSQ': [nngas_designcap,nngas_opcap,nngas_tsq,nngas_opavail,ngpl_designcap,ngpl_opcap,ngpl_tsq,ngpl_opavailcap,wt1_opcap,wt1_tsq,wt1_opavailcap,caprock_designcap,caprock_opcap,caprock_tsq,caprock_opavailcap,permn_designcap,permn_opcap,permn_tsq,permn_opavailcap,cornhpw_designcap,cornhpw_opcap,cornhpw_tsq,cornhpw_opavailcap,cornlpw_designcap,cornlpw_opcap,cornlpw_tsq,cornlpw_opavailcap,corntrn_designcap,corntrn_opcap,corntrn_tsq,corntrn_opavailcap,cornl2k_designcap,cornl2k_opcap,cornl2k_tsq,cornl2k_opavailcap]}
# Above dictionary has zipped the values of LOC to TSQ into a structure readable by Pandas to create a DataFrame with neatly organized loc and tsq values
dframe = pd.DataFrame(values,columns = ['LOC','TSQ']) # Creates a DataFrame
print(dframe) # Prints DataFrame
dframe.to_excel('SueDataBurner.xlsx') # Creates a burner Excel sheet with dataframe appended


# Below is the Excel Automation
burner = openpyxl.load_workbook("/Users/rajgarkhedkar/Desktop/Enkon/2020/SueAutomation/SueDataBurner.xlsx") # assigns this file to variable 'burner'
burn = burner['Sheet1'] # Assigns the proper sheet from burner workbook
pipedatabase = openpyxl.load_workbook('/Users/rajgarkhedkar/Desktop/Enkon/2020/SueAutomation/Sue Pipeline Data.xlsx') # Main file assigned to variable 'feedgas'

nngas_sheet = pipedatabase['Northern Natural Gas Co'] # Assigns this variable 'Northern Natural Gas Co' sheet on the Excel database
ngpl_sheet = pipedatabase['NGPL'] # Assigns this variable 'NGPL' sheet on the Excel database
transwestern_sheet = pipedatabase['Transwestern '] # Assigns this variable 'Transwestern' sheet on the Excel database
elpaso_sheet = pipedatabase['El Paso'] # Assigns this variable 'El Paso' sheet on the Excel database

brownfield_design = int(burn.cell(column=3,row=2).value)
brownfield_opcap = int(burn.cell(column=3,row=3).value)
brownfield_tsq = int(burn.cell(column=3,row=4).value)
brownfield_opavail = int(burn.cell(column=3,row=5).value)

north167_design = int(burn.cell(column=3,row=6).value)
north167_opcap = int(burn.cell(column=3,row=7).value)
north167_tsq = int(burn.cell(column=3,row=8).value)
north167_opavail = int(burn.cell(column=3,row=9).value)

wt1system_opcap = int(burn.cell(column=3,row=10).value)
wt1system_tsq = int(burn.cell(column=3,row=11).value)
wt1system_opavail = int(burn.cell(column=3,row=12).value)

caprockn_designcap = int(burn.cell(column=3,row=13).value.replace(',',''))
caprockn_opcap = int(burn.cell(column=3,row=14).value.replace(',',''))
caprockn_tsq = int(burn.cell(column=3,row=15).value.replace(',',''))
caprockn_opavail = int(burn.cell(column=3,row=16).value.replace(',',''))

perm_designcap = int(burn.cell(column=3,row=17).value.replace(',',''))
perm_opcap = int(burn.cell(column=3,row=18).value.replace(',',''))
perm_tsq = int(burn.cell(column=3,row=19).value.replace(',',''))
perm_opavail = int(burn.cell(column=3,row=20).value.replace(',',''))

cornhpwdesign = int(burn.cell(column=3,row=21).value.replace(',',''))
cornhpwopcap = int(burn.cell(column=3,row=22).value.replace(',',''))
cornhpwtsq = int(burn.cell(column=3,row=23).value.replace(',',''))
cornhpwopavail = int(burn.cell(column=3,row=24).value.replace(',',''))

cornlpwdesign = int(burn.cell(column=3,row=25).value.replace(',',''))
cornlpwopcap = int(burn.cell(column=3,row=26).value.replace(',',''))
cornlpwtsq = int(burn.cell(column=3,row=27).value.replace(',',''))
cornlpwopavail = int(burn.cell(column=3,row=28).value.replace(',',''))

corntrndesign = int(burn.cell(column=3,row=29).value.replace(',',''))
corntrnopcap = int(burn.cell(column=3,row=30).value.replace(',',''))
corntrntsq = int(burn.cell(column=3,row=31).value.replace(',',''))
corntrnopavail = int(burn.cell(column=3,row=32).value.replace(',',''))

cornl2kdesign = int(burn.cell(column=3,row=33).value.replace(',',''))
cornl2kopcap = int(burn.cell(column=3,row=34).value.replace(',',''))
cornl2ktsq = int(burn.cell(column=3,row=35).value.replace(',',''))
cornl2kopavail = int(burn.cell(column=3,row=36).value.replace(',',''))


nngas_sheet['G' + suenngcounter] = brownfield_design
nngas_sheet['H' + suenngcounter] = brownfield_opcap
nngas_sheet['I' + suenngcounter] = brownfield_tsq

ngpl_sheet['E' + suengplcounter] = north167_design
ngpl_sheet['F' + suengplcounter] = north167_opcap
ngpl_sheet['G' + suengplcounter] = north167_tsq

transwestern_sheet['F' + transcount] = wt1system_opcap
transwestern_sheet['G' + transcount] = wt1system_tsq

elpaso_sheet['E' + elpasocount] = caprockn_designcap
elpaso_sheet['F' + elpasocount] = caprockn_opcap
elpaso_sheet['G' + elpasocount] = caprockn_tsq

elpaso_sheet['L' + elpasocount] = perm_designcap
elpaso_sheet['M' + elpasocount] = perm_opcap
elpaso_sheet['N' + elpasocount] = perm_tsq

elpaso_sheet['S' + elpasocount] = cornhpwdesign
elpaso_sheet['T' + elpasocount] = cornhpwopcap
elpaso_sheet['U' + elpasocount] = cornhpwtsq

elpaso_sheet['X' + elpasocount] = cornlpwdesign
elpaso_sheet['Y' + elpasocount] = cornlpwopcap
elpaso_sheet['Z' + elpasocount] = cornlpwtsq

elpaso_sheet['AC' + elpasocount] = corntrndesign
elpaso_sheet['AD' + elpasocount] = corntrnopcap
elpaso_sheet['AE' + elpasocount] = corntrntsq

elpaso_sheet['AH' + elpasocount] = cornl2kdesign
elpaso_sheet['AI' + elpasocount] = cornl2kopcap
elpaso_sheet['AJ' + elpasocount] = cornl2ktsq

pipedatabase.save(filename='/Users/rajgarkhedkar/Desktop/Enkon/2020/SueAutomation/Sue Pipeline Data.xlsx')
os.remove('SueDataBurner.xlsx')

print("Daily Scraping Complete!")
