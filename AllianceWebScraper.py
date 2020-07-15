from bs4 import BeautifulSoup as soup
import requests
import pandas as pd
import re
import openpyxl
from pprint import pprint
import os


url = 'http://ips.alliance-pipeline.com/Ips/MainPage.aspx?siteCd=ALLUSA-IPS&contentSysCd=USA-OP-AVAIL-BY-DAY&tvPath=55/112/56'
html = soup(requests.get(url).content, 'html.parser')
loc = html.select('td.ig162a1706') # Specific cell 'td' and class 'ig162a1706' common for all loc cells
tsq = html.select('td.ig162a170e') # Specific cell 'td' and class 'ig162a170e' common for all tsq cells

data = {} # Empty dictionary to store the loc and tsq values
for location, quant in zip(loc, tsq): # Merges the values together, ties loc with tsq
    data[location.text] = quant.text # assigns a 1:1 loc to tsq relationship and appends to the dictionary

pprint(data) # Pretty prints the data

# Below is the Excel automation using pandas

df = pd.DataFrame(data=data, index=[0]) # Creates an empty dataframe and assigns the data values in a readable format for Excel
df = (df.T)  # Transposes index and column
print(df) # Prints output so user can ensure data validity
df.to_excel('AllianceBurner.xlsx') # Creates an excel file and appends all data into a column for loc and column for tsq. Unhash to use.



burner = openpyxl.load_workbook('/Users/rajgarkhedkar/Dropbox/Intern/AllianceBurner.xlsx') # Filepath for Excel data
burner_sheet = burner['Sheet1']
alliance_wb = openpyxl.load_workbook('/Users/rajgarkhedkar/Dropbox/Intern/Alliance Data 1Q 20.xlsx') # Filepath for Excel data
alliance_sheet = alliance_wb['Aux Sable']

alliance_anr = burner_sheet.cell(column=2, row=2).value
alliance_rosholt = burner_sheet.cell(column=2, row=3).value
aux_sable = burner_sheet.cell(column=2, row=4).value
bantry = burner_sheet.cell(column=2, row=5).value
border_usa = burner_sheet.cell(column=2, row=6).value
guardian = burner_sheet.cell(column=2, row=7).value
hankinson = burner_sheet.cell(column=2, row=8).value
horizon = burner_sheet.cell(column=2, row=9).value
lyle = burner_sheet.cell(column=2, row=10).value
midwestern_gas_transmission = burner_sheet.cell(column=2, row=11).value
milnor = burner_sheet.cell(column=2, row=12).value
natural_gas_pipeline_company = burner_sheet.cell(column=2, row=13).value
nicor_morris = burner_sheet.cell(column=2, row=14).value
peoples_elwood = burner_sheet.cell(column=2, row=15).value
tioga = burner_sheet.cell(column=2, row=16).value
vector = burner_sheet.cell(column=2, row=17).value

alliance_sheet['B1666'] = border_usa
alliance_sheet['D1666'] = tioga
alliance_sheet['F1666'] = bantry
alliance_sheet['H1666'] = hankinson
alliance_sheet['J1666'] = alliance_rosholt
alliance_sheet['L1666'] = lyle
alliance_sheet['N1666'] = milnor
alliance_sheet['Q1666'] = alliance_anr
alliance_sheet['S1666'] = aux_sable
alliance_sheet['U1666'] = guardian
alliance_sheet['W1666'] = midwestern_gas_transmission
alliance_sheet['Y1666'] = natural_gas_pipeline_company
alliance_sheet['AA1666'] = nicor_morris
alliance_sheet['AC1666'] = peoples_elwood
alliance_sheet['AE1666'] = vector
alliance_sheet['AG1666'] = horizon

alliance_wb.save(filename='/Users/rajgarkhedkar/Dropbox/Intern/Alliance Data 1Q 20.xlsx') # Unhash to use. Appends information to main file.

os.remove('AllianceBurner.xlsx')
print("Daily Web Scraping Data Complete.") 




#print(type(html))
#print(html.name)
#print(html.prettify()) # This prints out the entire html from the page in order to find specific table elements
#print(html.body.prettify()) # This prints out just the body of the html page in order to find specific table elements



